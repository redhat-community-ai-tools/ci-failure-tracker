"""
Flask web server for dashboard
"""

from flask import Flask, render_template, jsonify, request, send_file, session, redirect
from datetime import datetime, timedelta
from pathlib import Path
import time
import yaml
import threading
import sys
import os
import logging
import io
import csv
from collections import OrderedDict
import html as html_module
import secrets
from urllib.parse import urlencode
import requests as http_requests
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from storage.database import DashboardDatabase
from metrics.calculator import MetricsCalculator
from reports.weekly_report import WeeklyReportGenerator

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Global collection status
collection_status = {
    'running': False,
    'progress': '',
    'error': None,
    'completed_at': None,
    'lock': threading.Lock()
}

# ---------------------------------------------------------------------------
# Bounded token store – prevents unbounded memory growth from OAuth tokens
# that are never explicitly logged-out (e.g. user closes browser tab).
#
# Only a random token_id is placed in the (signed, client-side) Flask session
# cookie so the actual GitHub token is never exposed to the browser.
# Note: tokens are lost on process restart; users simply re-authenticate.
# ---------------------------------------------------------------------------

# Defaults – override via env vars TOKEN_STORE_MAX_SIZE / TOKEN_STORE_TTL_SECONDS
_TOKEN_STORE_MAX_SIZE = int(os.environ.get('TOKEN_STORE_MAX_SIZE', '1024'))
_TOKEN_STORE_TTL_SECONDS = int(
    os.environ.get('TOKEN_STORE_TTL_SECONDS', str(8 * 3600))   # 8 hours
)


class _BoundedTokenStore:
    """Dict-like store with TTL expiry and LRU eviction.

    Each entry records its insertion timestamp.  Expired entries are
    lazily purged on ``get`` and ``__setitem__``.  When the store
    reaches *max_size*, the oldest entry (by insertion order) is
    evicted regardless of its age.
    """

    def __init__(self, max_size=_TOKEN_STORE_MAX_SIZE,
                 max_age=_TOKEN_STORE_TTL_SECONDS, clock=time.time):
        self._data = OrderedDict()       # token_id -> (value, timestamp)
        self._max_size = max_size
        self._max_age = max_age
        self._clock = clock              # injectable for testing

    # -- internal helpers ---------------------------------------------------

    def _is_expired(self, timestamp):
        return (self._clock() - timestamp) > self._max_age

    def _evict_expired(self):
        """Remove all entries older than *max_age*."""
        now = self._clock()
        # OrderedDict is insertion-ordered; oldest entries are at the front.
        keys_to_delete = [
            k for k, (_, ts) in self._data.items()
            if (now - ts) > self._max_age
        ]
        for k in keys_to_delete:
            del self._data[k]

    # -- dict-compatible public API -----------------------------------------

    def __setitem__(self, key, value):
        self._evict_expired()
        # Remove first so re-insertion moves it to the end (most-recent).
        if key in self._data:
            del self._data[key]
        self._data[key] = (value, self._clock())
        # Enforce max-size: evict oldest entries.
        while len(self._data) > self._max_size:
            self._data.popitem(last=False)

    def get(self, key, default=None):
        entry = self._data.get(key)
        if entry is None:
            return default
        value, ts = entry
        if self._is_expired(ts):
            del self._data[key]
            return default
        return value

    def pop(self, key, *args):
        entry = self._data.pop(key, None)
        if entry is None:
            if args:
                return args[0]
            raise KeyError(key)
        return entry[0]

    def clear(self):
        self._data.clear()

    def __contains__(self, key):
        entry = self._data.get(key)
        if entry is None:
            return False
        if self._is_expired(entry[1]):
            del self._data[key]
            return False
        return True

    def __len__(self):
        self._evict_expired()
        return len(self._data)

    def values(self):
        self._evict_expired()
        return [v for v, _ in self._data.values()]


_oauth_token_store = _BoundedTokenStore()


def run_collection_background(db_path: str, config_file: str = 'config.yaml', days: int = 30, version_filter: str = ''):
    """Run data collection in background thread"""
    global collection_status

    try:
        logger.info(f"Starting data collection for {days} days" +
                     (f", version={version_filter}" if version_filter else ""))
        collection_status['progress'] = 'Starting collection...'

        # Load config
        with open(config_file, 'r') as f:
            config = yaml.safe_load(f)

        # Import collector modules
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

        # Initialize collector based on type
        collector_type = config['collector']['type']
        logger.info(f"Using collector type: {collector_type}")

        # Shared tracking config injected into collector config
        tracking = config.get('tracking', {})
        test_suite_filter = tracking.get('test_suite_filter', '')
        branch_version_map = tracking.get('branch_version_map', {})

        if collector_type == 'reportportal':
            from collectors.reportportal import ReportPortalCollector
            rp_config = config['collector']['reportportal'].copy()
            rp_config['test_suite_filter'] = test_suite_filter
            rp_config['branch_version_map'] = branch_version_map
            collector = ReportPortalCollector(rp_config)
        elif collector_type == 'prow_mcp':
            from collectors.prow_mcp import ProwMCPCollector
            mcp_config = config['collector']['prow_mcp'].copy()
            mcp_config['branch_version_map'] = branch_version_map
            collector = ProwMCPCollector(mcp_config)
        elif collector_type == 'prow_gcs':
            from collectors.prow_gcs import ProwGCSCollector
            gcs_config = config['collector']['prow_gcs'].copy()
            gcs_config['test_suite_filter'] = test_suite_filter
            gcs_config['branch_version_map'] = branch_version_map
            try:
                collector = ProwGCSCollector(gcs_config)
            except Exception as e:
                error_msg = f'Failed to initialize prow_gcs collector: {e}'
                logger.error(error_msg)
                collection_status['error'] = error_msg
                collection_status['running'] = False
                return
        elif collector_type == 'gcsweb':
            from collectors.gcsweb import GCSWebCollector
            gcsweb_config = config['collector']['gcsweb'].copy()
            gcsweb_config['test_suite_filter'] = test_suite_filter
            gcsweb_config['branch_version_map'] = branch_version_map
            collector = GCSWebCollector(gcsweb_config)
        else:
            error_msg = f'Unsupported collector type: {collector_type}'
            logger.error(error_msg)
            collection_status['error'] = error_msg
            collection_status['running'] = False
            return

        # Health check
        logger.info("Running health check...")
        collection_status['progress'] = 'Checking data source...'
        if not collector.health_check():
            error_msg = getattr(collector, 'health_error', None) or 'Failed to connect to data source'
            logger.error(error_msg)
            collection_status['error'] = error_msg
            collection_status['running'] = False
            return

        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

        # Get job patterns based on collector type
        versions = config['tracking']['versions']
        platforms = config['tracking']['platforms']

        # Apply version filter from user's selection
        if version_filter:
            versions = [v for v in versions if v == version_filter]
            if not versions:
                versions = [version_filter]

        if collector_type == 'reportportal':
            job_patterns = config['collector']['reportportal']['job_patterns']
            # Expand patterns with version placeholders
            expanded_patterns = []
            for pattern in job_patterns:
                for version in versions:
                    expanded_patterns.append(pattern.replace('{version}', version))
        elif collector_type == 'prow_gcs':
            # prow_gcs uses wildcard patterns, no version expansion needed
            # Support both 'job_patterns' (new) and 'job_names' (legacy)
            prow_gcs_config = config['collector']['prow_gcs']
            expanded_patterns = prow_gcs_config.get('job_patterns') or prow_gcs_config.get('job_names', [])
        elif collector_type == 'prow_mcp':
            # prow_mcp uses exact job names from config
            expanded_patterns = None  # Will use job_names from collector config
        elif collector_type == 'gcsweb':
            gcsweb_cfg = config['collector']['gcsweb']
            all_job_names = list(gcsweb_cfg['job_names'])
            # Include postsubmit job patterns if configured
            postsubmit_patterns = gcsweb_cfg.get('postsubmit_job_patterns', [])
            all_job_names.extend(postsubmit_patterns)
            if version_filter:
                branch_map = config.get('tracking', {}).get('branch_version_map', {})
                reverse_map = {v: k for k, v in branch_map.items()}
                branch = reverse_map.get(version_filter)
                expanded_patterns = [
                    j for j in all_job_names
                    if f'release-{version_filter}' in j or (branch and f'-{branch}-' in j)
                ]
                if not expanded_patterns:
                    expanded_patterns = all_job_names
            else:
                expanded_patterns = all_job_names
        else:
            expanded_patterns = []

        # Use single-pass incremental collection for gcsweb
        if collector_type == 'gcsweb' and hasattr(collector, 'collect_all'):
            db = DashboardDatabase(db_path)
            skip_builds = db.get_existing_build_ids(expanded_patterns)
            logger.info(f"Incremental collection: {len(skip_builds)} builds already in DB")
            collection_status['progress'] = f'Found {len(skip_builds)} existing builds, collecting new...'

            def _progress(msg):
                collection_status['progress'] = msg

            # Get PR log sources config (defaults to empty list)
            pr_log_sources = config['collector']['gcsweb'].get('pr_log_sources', [])

            job_runs, test_results = collector.collect_all(
                start_date=start_date,
                end_date=end_date,
                job_patterns=expanded_patterns,
                versions=versions,
                platforms=platforms,
                skip_builds=skip_builds,
                progress_callback=_progress,
                pr_log_sources=pr_log_sources
            )
            logger.info(f"Collected {len(job_runs)} job runs, {len(test_results)} test results (single pass)")
        else:
            # Legacy two-pass collection for other collector types
            logger.info("Collecting job runs...")
            collection_status['progress'] = 'Collecting job runs...'
            job_runs = collector.collect_job_runs(
                start_date=start_date,
                end_date=end_date,
                job_patterns=expanded_patterns,
                versions=versions,
                platforms=platforms
            )
            logger.info(f"Collected {len(job_runs)} job runs")

            collection_status['progress'] = f'Collected {len(job_runs)} job runs, collecting test results...'
            logger.info("Collecting test results...")
            test_results = collector.collect_test_results(
                start_date=start_date,
                end_date=end_date,
                job_patterns=expanded_patterns,
                versions=versions,
                platforms=platforms
            )
            logger.info(f"Collected {len(test_results)} test results")

            db = DashboardDatabase(db_path)

        # Save to database
        collection_status['progress'] = f'Saving {len(job_runs)} job runs, {len(test_results)} test results...'
        logger.info("Saving to database...")

        inserted_jobs = db.insert_job_runs(job_runs)
        inserted_tests = db.insert_test_results(test_results)

        # Update job_runs with actual test counts from test_results
        logger.info("Updating job runs with test counts...")
        db.conn.execute("""
            UPDATE job_runs
            SET
                total_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status != 'skipped'
                ),
                passed_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'passed'
                ),
                failed_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'failed'
                ),
                skipped_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'skipped'
                )
            WHERE EXISTS (
                SELECT 1 FROM test_results
                WHERE test_results.job_name = job_runs.job_name
                AND test_results.build_id = job_runs.build_id
            )
        """)
        db.conn.commit()
        logger.info("Job runs updated with test counts")

        # Close connection after write
        db.conn.close()

        db.close()

        logger.info(f"Collection complete! Inserted {inserted_jobs} job runs and {inserted_tests} test results")
        collection_status['progress'] = f'Complete! Saved {inserted_jobs} job runs and {inserted_tests} test results'
        collection_status['error'] = None
        collection_status['completed_at'] = datetime.now().isoformat()

    except Exception as e:
        logger.error(f"Collection failed: {e}", exc_info=True)
        collection_status['error'] = str(e)
        collection_status['progress'] = 'Failed'
        collection_status['completed_at'] = None
    finally:
        logger.info("Collection thread finished")
        collection_status['running'] = False


def create_app(db_path: str, config: dict = None, config_file: str = 'config.yaml'):
    """
    Create Flask application

    Args:
        db_path: Path to SQLite database
        config: Optional Flask configuration
        config_file: Path to YAML configuration file

    Returns:
        Flask app instance
    """
    app = Flask(__name__,
                template_folder=str(Path(__file__).parent / 'templates'),
                static_folder=str(Path(__file__).parent / 'static'))

    # Disable template caching for development
    app.config['TEMPLATES_AUTO_RELOAD'] = True
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.jinja_env.auto_reload = True
    app.jinja_env.cache = {}

    # Configure session secret key for OAuth flow
    app.secret_key = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))

    if config:
        app.config.update(config)

    # GitHub OAuth configuration
    github_oauth_client_id = os.environ.get('GITHUB_OAUTH_CLIENT_ID', '')
    github_oauth_client_secret = os.environ.get('GITHUB_OAUTH_CLIENT_SECRET', '')

    # Warn if OAuth is configured but secret key is ephemeral
    if (github_oauth_client_id and github_oauth_client_secret
            and not os.environ.get('FLASK_SECRET_KEY')):
        logger.warning(
            "FLASK_SECRET_KEY is not set. OAuth sessions will not survive "
            "pod restarts. Set FLASK_SECRET_KEY to a stable value for "
            "reliable OAuth."
        )

    # Load tracking config for blocklist and configured versions/platforms
    blocklist = []
    config_versions = []
    config_platforms = []
    try:
        with open(config_file, 'r') as f:
            yaml_config = yaml.safe_load(f)
            blocklist = yaml_config.get('tracking', {}).get('blocklist', [])
            config_versions = yaml_config.get('tracking', {}).get('versions', [])
            config_platforms = yaml_config.get('tracking', {}).get('platforms', [])
    except Exception as e:
        print(f"Warning: Could not load tracking config: {e}")

    # Initialize database and calculator
    db = DashboardDatabase(db_path)
    calculator = MetricsCalculator(db, blocklist=blocklist)
    report_generator = WeeklyReportGenerator(db, blocklist=blocklist)

    # Check if AI analysis is enabled (default: False for production safety)
    enable_ai = os.environ.get('ENABLE_AI_ANALYSIS', 'false').lower() == 'true'

    def get_latest_version():
        """
        Get the latest version from database.
        Returns the highest version number (e.g., "4.22" if both "4.21" and "4.22" exist)
        """
        query = "SELECT DISTINCT version FROM job_runs ORDER BY version DESC LIMIT 1"
        result = db.execute_query(query)
        return result[0]['version'] if result else None

    def normalize_version(version):
        """
        Normalize version parameter: if empty/None, return latest version.
        This prevents statistically invalid aggregation across different versions.
        """
        if not version or version == '':
            return get_latest_version()
        return version

    @app.route('/')
    def index():
        """Render main dashboard page.
        Shows whatever data is already in the DB from cron/manual collection.
        Users trigger collection manually via the refresh button."""
        github_repo = os.environ.get('GITHUB_REPO', '')
        notify_users_raw = os.environ.get('GITHUB_NOTIFY_USERS', '')
        github_notify_users = [
            u.strip().lstrip('@')
            for u in notify_users_raw.split(',')
            if u.strip()
        ]
        return render_template(
            'dashboard.html',
            enable_ai=enable_ai,
            github_repo=github_repo,
            github_notify_users=github_notify_users,
        )

    @app.route('/logs')
    def view_logs():
        """Display test logs in a new page"""
        log_content = html_module.escape(request.args.get('content', ''))
        test_name = html_module.escape(request.args.get('test', 'Test Log'))

        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>{test_name} - Logs</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: #f8fafc;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                .header {{
                    background: #1e40af;
                    color: white;
                    padding: 20px;
                    font-size: 18px;
                    font-weight: 600;
                }}
                .content {{
                    padding: 20px;
                }}
                pre {{
                    background: #1e293b;
                    color: #e2e8f0;
                    padding: 20px;
                    border-radius: 6px;
                    overflow-x: auto;
                    font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                    font-size: 13px;
                    line-height: 1.6;
                    white-space: pre-wrap;
                    word-wrap: break-word;
                }}
                .error {{
                    color: #fca5a5;
                }}
                .info {{
                    color: #93c5fd;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">{test_name}</div>
                <div class="content">
                    <pre>{log_content}</pre>
                </div>
            </div>
        </body>
        </html>
        '''
        return html

    @app.route('/api/collection-status')
    def api_collection_status():
        """Get current collection status"""
        global collection_status
        return jsonify({
            'running': collection_status['running'],
            'progress': collection_status['progress'],
            'error': collection_status['error'],
            'completed_at': collection_status['completed_at']
        })

    @app.route('/api/trigger-collection', methods=['POST'])
    def api_trigger_collection():
        """Manually trigger data collection"""
        global collection_status

        data = request.json or {}
        days = data.get('days', 30)
        version = data.get('version', '')

        with collection_status['lock']:
            if collection_status['running']:
                return jsonify({'error': 'Collection already running'}), 409

            collection_status['running'] = True
            collection_status['progress'] = 'Initializing...'
            collection_status['error'] = None
            collection_status['completed_at'] = None

            thread = threading.Thread(
                target=run_collection_background,
                args=(db_path, config_file, days, version),
                daemon=True
            )
            thread.start()

        return jsonify({'status': 'started'})

    @app.route('/api/metadata')
    def api_metadata():
        """Get available versions and platforms from database + config"""
        query_versions = "SELECT DISTINCT version FROM job_runs ORDER BY version DESC"
        query_platforms = "SELECT DISTINCT platform FROM job_runs ORDER BY platform"

        db_versions = [row['version'] for row in db.execute_query(query_versions)]
        db_platforms = [row['platform'] for row in db.execute_query(query_platforms)]

        all_versions = sorted(
            set(db_versions + config_versions),
            key=lambda v: [int(x) for x in v.split('.')],
            reverse=True
        )
        all_platforms = sorted(set(db_platforms + config_platforms))

        return jsonify({
            'versions': all_versions,
            'platforms': all_platforms
        })

    @app.route('/api/summary')
    def api_summary():
        """Get summary statistics"""
        days = request.args.get('days', 7, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        stats = calculator.get_summary_stats(days=days, version=version, platform=platform)
        return jsonify(stats)

    @app.route('/api/trend')
    def api_trend():
        """Get overall pass rate trend"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')

        trend = calculator.get_overall_trend(
            days=days,
            version=version,
            platform=platform
        )
        return jsonify(trend)

    @app.route('/api/test-rankings')
    def api_test_rankings():
        """Get test rankings (worst performers)"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        limit = request.args.get('limit', 20, type=int)

        rankings = calculator.get_test_rankings(
            days=days,
            version=version,
            platform=platform,
            limit=limit
        )
        return jsonify(rankings)

    @app.route('/api/version-comparison')
    def api_version_comparison():
        """Compare pass rates across versions"""
        days = request.args.get('days', 30, type=int)
        comparison = calculator.get_version_comparison(days=days)
        return jsonify(comparison)

    @app.route('/api/platform-comparison')
    def api_platform_comparison():
        """Compare pass rates across platforms"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))

        comparison = calculator.get_platform_comparison(
            days=days,
            version=version
        )
        return jsonify(comparison)

    @app.route('/api/weekly-report')
    def api_weekly_report():
        """Get weekly platform breakdown report"""
        current_days = request.args.get('current_days', 7, type=int)
        previous_days = request.args.get('previous_days', 7, type=int)
        version = normalize_version(request.args.get('version'))
        top = request.args.get('top', 10, type=int)

        # Get platform comparison
        comparison = report_generator.get_platform_week_over_week(
            current_week_days=current_days,
            previous_week_days=previous_days,
            version=version
        )

        # Get top failing tests
        top_tests = calculator.get_test_rankings(days=current_days, version=version, limit=top)

        # Get overall summary
        summary = calculator.get_summary_stats(days=current_days, version=version)

        return jsonify({
            'comparison': comparison,
            'top_tests': top_tests,
            'summary': summary
        })

    @app.route('/api/build-health')
    def api_build_health():
        """Get build health summary grouped by operator (WMCO) version.

        Returns the latest operator version with per-platform pass/fail
        breakdown.  Only includes job runs where the operator version was
        successfully extracted from build logs.
        """
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))

        rows = db.get_build_health(version=version, days=days)

        if not rows:
            return jsonify({
                'operator_versions': [],
                'latest_version': None,
                'platforms': {},
            })

        # Group by operator_version
        versions_data = {}
        for row in rows:
            ov = row['operator_version']
            if ov not in versions_data:
                versions_data[ov] = {'platforms': {}, 'total_runs': 0,
                                     'passed_runs': 0, 'failed_runs': 0}
            versions_data[ov]['platforms'][row['platform']] = {
                'total_runs': row['total_runs'],
                'passed_runs': row['passed_runs'],
                'failed_runs': row['failed_runs'],
            }
            versions_data[ov]['total_runs'] += row['total_runs']
            versions_data[ov]['passed_runs'] += row['passed_runs']
            versions_data[ov]['failed_runs'] += row['failed_runs']

        # Sort versions descending; the "latest" is the first one
        sorted_versions = sorted(versions_data.keys(), reverse=True)
        latest = sorted_versions[0] if sorted_versions else None

        # Compute releasable flag: all platforms passed every run
        for ov, data in versions_data.items():
            data['releasable'] = all(
                p['failed_runs'] == 0
                for p in data['platforms'].values()
            )
            data['pass_rate'] = round(
                data['passed_runs'] / data['total_runs'] * 100, 1
            ) if data['total_runs'] else 0.0

        # Only return latest version per the feature request
        result = {
            'latest_version': latest,
            'platforms': versions_data.get(latest, {}).get('platforms', {}),
            'total_runs': versions_data.get(latest, {}).get('total_runs', 0),
            'passed_runs': versions_data.get(latest, {}).get('passed_runs', 0),
            'failed_runs': versions_data.get(latest, {}).get('failed_runs', 0),
            'pass_rate': versions_data.get(latest, {}).get('pass_rate', 0.0),
            'releasable': versions_data.get(latest, {}).get('releasable', False),
        }

        # Add GitHub source link if commit hash is present
        if latest and '-' in latest:
            commit_hash = latest.split('-', 1)[1]
            result['source_url'] = (
                f'https://github.com/openshift/windows-machine-config-operator'
                f'/commit/{commit_hash}'
            )

        return jsonify(result)

    @app.route('/api/platform-tests')
    def api_platform_tests():
        """Get test results for a specific platform"""
        platform = request.args.get('platform')
        days = request.args.get('days', 7, type=int)
        version = normalize_version(request.args.get('version'))

        if not platform:
            return jsonify({'error': 'Platform parameter is required'}), 400

        # Get test rankings for this platform
        tests = calculator.get_test_rankings(days=days, version=version, platform=platform, limit=100)

        # Get platform-specific summary
        summary = calculator.get_summary_stats(days=days, platform=platform, version=version)

        return jsonify({
            'platform': platform,
            'tests': tests,
            'summary': summary,
            'days': days
        })

    @app.route('/api/test-error-by-platform')
    def api_test_error_by_platform():
        """Get latest error for a specific test on a specific platform"""
        test_name = request.args.get('test_name')
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        days = request.args.get('days', 30, type=int)

        if not test_name or not platform:
            return jsonify({'error': 'test_name and platform parameters are required'}), 400

        # Query for most recent failure on this platform
        query = """
            SELECT
                error_message,
                timestamp,
                job_name,
                build_id,
                job_url,
                platform,
                log_url
            FROM test_results
            WHERE test_name = ?
            AND platform = ?
            AND status = 'failed'
            AND error_message IS NOT NULL
            AND timestamp >= datetime('now', ? || ' days')
        """

        params = [test_name, platform, f'-{days}']

        if version:
            query += " AND version = ?"
            params.append(version)

        query += " ORDER BY timestamp DESC LIMIT 1"

        result = db.execute_query(query, params)

        if result:
            return jsonify(result[0])
        else:
            return jsonify({'error': 'No error found for this test/platform combination'}), 404

    @app.route('/api/get-affected-platforms', methods=['POST'])
    def api_get_affected_platforms():
        """Get all platforms affected by a test failure"""
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        days = data.get('days', 7)

        if not all([test_name, version]):
            return jsonify({'error': 'Missing required fields: test_name, version'}), 400

        platforms = db.get_affected_platforms(test_name, version, days)
        return jsonify({'platforms': platforms})

    @app.route('/api/jira/create', methods=['POST'])
    def api_create_jira():
        """Create or find existing Jira issue for a test failure"""
        from integrations import get_jira_integration

        jira = get_jira_integration()
        if not jira:
            return jsonify({
                'status': 'disabled',
                'message': 'Jira integration not configured. Set JIRA_API_TOKEN environment variable.'
            })

        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        # Required fields
        test_name = data.get('test_name')
        version = data.get('version')
        platforms = data.get('platforms', [])

        if not all([test_name, version]):
            return jsonify({'error': 'Missing required fields: test_name, version'}), 400

        # If no platforms provided, use single platform from old API
        if not platforms:
            platform = data.get('platform')
            if platform:
                platforms = [platform]

        # Optional fields
        test_description = data.get('test_description', '')
        error_message = data.get('error_message', '')
        job_url = data.get('job_url', '')
        failure_rate = data.get('failure_rate', 0.0)
        runs = data.get('runs', 0)
        failures = data.get('failures', 0)

        # Check for existing issue first (search by test_name + version only)
        existing_issue = jira.search_existing_issue(test_name, version)
        if existing_issue:
            issue_key = existing_issue.get('key')
            issue_url = jira.get_issue_url(issue_key)
            # Save to database (applies to all platforms)
            db.save_jira_issue(test_name, version, jira_issue_key=issue_key)
            return jsonify({
                'status': 'existing',
                'issue_key': issue_key,
                'issue_url': issue_url,
                'message': f'Found existing issue: {issue_key}'
            })

        # Create new issue
        issue_key = jira.create_issue(
            test_name=test_name,
            test_description=test_description,
            version=version,
            platforms=platforms,
            error_message=error_message,
            job_url=job_url,
            failure_rate=failure_rate,
            runs=runs,
            failures=failures
        )

        if issue_key:
            issue_url = jira.get_issue_url(issue_key)
            # Save to database (applies to all platforms)
            db.save_jira_issue(test_name, version, jira_issue_key=issue_key)
            return jsonify({
                'status': 'created',
                'issue_key': issue_key,
                'issue_url': issue_url,
                'message': f'Created new issue: {issue_key}'
            })
        else:
            return jsonify({'error': 'Failed to create Jira issue'}), 500

    @app.route('/auth/github/login')
    def auth_github_login():
        """Redirect user to GitHub OAuth authorization page."""
        if not github_oauth_client_id or not github_oauth_client_secret:
            return jsonify({'error': 'GitHub OAuth is not configured'}), 400

        # Generate a random state parameter to prevent CSRF
        state = secrets.token_hex(16)
        session['oauth_state'] = state

        params = {
            'client_id': github_oauth_client_id,
            'redirect_uri': request.url_root.rstrip('/') + '/auth/github/callback',
            'scope': 'public_repo',
            'state': state,
        }
        return redirect(f'https://github.com/login/oauth/authorize?{urlencode(params)}')

    @app.route('/auth/github/callback')
    def auth_github_callback():
        """Handle GitHub OAuth callback and exchange code for token."""
        if not github_oauth_client_id or not github_oauth_client_secret:
            return jsonify({'error': 'GitHub OAuth is not configured'}), 400

        code = request.args.get('code')
        state = request.args.get('state')

        if not code:
            return jsonify({'error': 'Missing authorization code'}), 400

        # Verify state parameter to prevent CSRF
        stored_state = session.pop('oauth_state', None)
        if not state or state != stored_state:
            return jsonify({'error': 'Invalid state parameter'}), 400

        # Exchange code for access token
        try:
            token_response = http_requests.post(
                'https://github.com/login/oauth/access_token',
                json={
                    'client_id': github_oauth_client_id,
                    'client_secret': github_oauth_client_secret,
                    'code': code,
                },
                headers={'Accept': 'application/json'},
                timeout=30,
            )

            if token_response.status_code != 200:
                logger.error(f"GitHub token exchange failed: {token_response.status_code}")
                return jsonify({'error': 'Failed to exchange authorization code'}), 500

            token_data = token_response.json()
            access_token = token_data.get('access_token')

            if not access_token:
                error_desc = token_data.get('error_description', 'Unknown error')
                logger.error(f"GitHub OAuth error: {error_desc}")
                return jsonify({'error': f'OAuth error: {error_desc}'}), 400

            # Fetch the user's GitHub profile to get their username
            user_response = http_requests.get(
                'https://api.github.com/user',
                headers={
                    'Authorization': f'token {access_token}',
                    'Accept': 'application/vnd.github.v3+json',
                },
                timeout=30,
            )

            if user_response.status_code != 200:
                logger.error(f"GitHub user fetch failed: {user_response.status_code}")
                return jsonify({'error': 'Failed to fetch GitHub user info'}), 500

            user_data = user_response.json()
            github_username = user_data.get('login', '')

            # Store access token server-side; only a random id goes into
            # the signed (not encrypted) session cookie.
            token_id = secrets.token_hex(16)
            _oauth_token_store[token_id] = access_token
            session['oauth_token_id'] = token_id
            session['github_username'] = github_username

            logger.info(f"GitHub OAuth login successful for user: {github_username}")

            # Redirect back to the dashboard
            return redirect('/')

        except Exception as e:
            logger.error(f"GitHub OAuth callback error: {e}")
            return jsonify({'error': 'OAuth callback failed'}), 500

    @app.route('/auth/github/status')
    def auth_github_status():
        """Return whether the user is authenticated via GitHub OAuth."""
        token_id = session.get('oauth_token_id')
        access_token = _oauth_token_store.get(token_id) if token_id else None
        username = session.get('github_username')
        oauth_configured = bool(github_oauth_client_id and github_oauth_client_secret)

        return jsonify({
            'authenticated': bool(access_token and username),
            'username': username or '',
            'oauth_configured': oauth_configured,
        })

    @app.route('/auth/github/logout', methods=['POST'])
    def auth_github_logout():
        """Clear the GitHub OAuth session."""
        token_id = session.pop('oauth_token_id', None)
        if token_id:
            _oauth_token_store.pop(token_id, None)
        session.pop('github_username', None)
        return jsonify({'status': 'logged_out'})

    @app.route('/api/analyze-failure', methods=['POST'])
    def api_analyze_failure():
        """
        Analyze test failure with AI (hybrid: local Claude Code or Anthropic API)
        """
        from ai.analyzer import HybridFailureAnalyzer

        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')

        if not all([test_name, version, platform]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform'}), 400

        # Check if we already have a recent analysis
        days = data.get('days', 7)
        existing_analysis = db.get_ai_analysis(test_name, version, platform, days)
        if existing_analysis and data.get('use_cached', True):
            existing_analysis['cached'] = True
            return jsonify(existing_analysis)

        # Use provided error_message or get from database
        error_message = data.get('error_message')
        log_url = data.get('log_url', '')
        test_description = data.get('test_description', '')

        if not error_message:
            # Get test error details from database
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)

            query = """
                SELECT error_message, log_url, test_description
                FROM test_results
                WHERE test_name = ?
                AND version = ?
                AND platform = ?
                AND status = 'failed'
                AND timestamp >= ? AND timestamp <= ?
                ORDER BY timestamp DESC
                LIMIT 1
            """

            cursor = db.conn.cursor()
            cursor.execute(query, (test_name, version, platform,
                                   start_date.isoformat(), end_date.isoformat()))
            test_data = cursor.fetchone()

            if not test_data:
                return jsonify({'error': 'No recent failure found for this test'}), 404

            error_message = test_data[0] or 'No error message'
            log_url = test_data[1] or ''
            test_description = test_data[2] or ''

        # Get pass rate for pre-classifier
        pass_rate = None
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            cursor = db.conn.cursor()
            cursor.execute("""
                SELECT CAST(SUM(CASE WHEN status = 'passed' THEN 1 ELSE 0 END) AS REAL)
                       / NULLIF(SUM(CASE WHEN status IN ('passed', 'failed') THEN 1 ELSE 0 END), 0) * 100
                FROM test_results
                WHERE test_name = ? AND version = ?
                AND status != 'skipped'
                AND timestamp >= ? AND timestamp <= ?
            """, (test_name, version, start_date.isoformat(), end_date.isoformat()))
            row = cursor.fetchone()
            if row and row[0] is not None:
                pass_rate = row[0]
        except Exception:
            pass

        # Analyze with pre-classifier + AI
        try:
            analyzer = HybridFailureAnalyzer()
            analysis = analyzer.analyze_failure(
                test_name=test_name,
                error_message=error_message,
                log_url=log_url,
                platform=platform,
                version=version,
                pass_rate=pass_rate,
                test_description=test_description
            )

            # Save analysis to database
            db.save_ai_analysis(test_name, version, platform, analysis)

            analysis['cached'] = False
            return jsonify(analysis)

        except Exception as e:
            return jsonify({
                'error': f'Analysis failed: {str(e)}',
                'root_cause': 'Analysis service error',
                'confidence': 0
            }), 500

    @app.route('/api/analysis-stats')
    def api_analysis_stats():
        """Get statistics about AI analyses"""
        stats = db.get_analysis_stats()
        return jsonify(stats)

    @app.route('/api/save-classification', methods=['POST'])
    def api_save_classification():
        """
        Save manual classification for a test failure
        """
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')
        classification = data.get('classification')

        if not all([test_name, version, platform, classification]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform, classification'}), 400

        # Validate classification
        valid_classifications = ['product_bug', 'automation_bug', 'system_issue', 'transient', 'to_investigate']
        if classification not in valid_classifications:
            return jsonify({'error': f'Invalid classification. Must be one of: {", ".join(valid_classifications)}'}), 400

        # Save to database
        rows_updated = db.save_manual_classification(
            test_name=test_name,
            version=version,
            platform=platform,
            classification=classification,
            classified_by='user'
        )

        if rows_updated > 0:
            return jsonify({
                'status': 'success',
                'rows_updated': rows_updated,
                'classification': classification
            })
        else:
            return jsonify({'error': 'No matching test result found to update'}), 404

    @app.route('/api/get-test-data', methods=['POST'])
    def api_get_test_data():
        """
        Get existing data for a test (classification, Jira key, AI analysis)
        """
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')

        if not all([test_name, version, platform]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform'}), 400

        result = {
            'manual_classification': None,
            'jira_issue_key': None,
            'ai_analysis': None
        }

        # Get manual classification and Jira issue from test_results
        cursor = db.conn.cursor()

        # Log query parameters for debugging
        logger.info(f"Fetching test data: test_name={test_name}, version={version}, platform={platform}")

        cursor.execute("""
            SELECT manual_classification, jira_issue_key
            FROM test_results
            WHERE test_name = ?
            AND version = ?
            AND UPPER(platform) = UPPER(?)
            AND status = 'failed'
            ORDER BY timestamp DESC
            LIMIT 1
        """, (test_name, version, platform))

        row = cursor.fetchone()
        if row:
            result['manual_classification'] = row[0]
            result['jira_issue_key'] = row[1]
            logger.info(f"Found test data: classification={row[0]}, jira_key={row[1]}")
        else:
            logger.info(f"No test data found for {test_name}/{version}/{platform}")

        # Get AI analysis
        ai_analysis = db.get_ai_analysis(test_name, version, platform, days=90)
        if ai_analysis:
            result['ai_analysis'] = ai_analysis

        return jsonify(result)

    @app.route('/api/export')
    def api_export():
        """Export test results to XLSX, CSV, or MD format"""
        export_format = request.args.get('format', 'xlsx')
        days = request.args.get('days', 30, type=int)
        version_param = request.args.get('version')
        version = normalize_version(version_param)

        # Debug logging
        logger.info(f"[EXPORT] Received: format={export_format}, days={days}, version_param={version_param}, normalized_version={version}")

        # Get metadata to get all platforms
        query = "SELECT DISTINCT platform FROM test_results WHERE platform IS NOT NULL ORDER BY platform"
        platforms_data = db.execute_query(query)
        platforms = [row['platform'] for row in platforms_data] if platforms_data else []

        # Collect data for all platforms
        all_data = {}
        pass_rates = {}

        for platform in platforms:
            logger.info(f"[EXPORT] Fetching {platform} data: days={days}, version={version}")
            tests = calculator.get_test_rankings(days=days, version=version, platform=platform, limit=1000)
            logger.info(f"[EXPORT] {platform}: Found {len(tests)} tests")
            all_data[platform] = tests

            # Calculate pass rate for this platform
            if tests:
                total_executions = sum(test['total_runs'] for test in tests)
                passed_executions = sum(test['passed_runs'] for test in tests)
                pass_rate = (passed_executions / total_executions * 100) if total_executions > 0 else 0
                pass_rates[platform] = pass_rate

        # Generate file based on format
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'dashboard-export-{version}-{days}days-{today}'

        if export_format == 'xlsx':
            return export_to_xlsx(all_data, pass_rates, filename, version, days)
        elif export_format == 'csv':
            return export_to_csv(all_data, filename, version, days)
        elif export_format == 'md':
            return export_to_markdown(all_data, filename, version, days)
        else:
            return jsonify({'error': 'Invalid format. Use xlsx, csv, or md'}), 400

    def export_to_xlsx(all_data, pass_rates, filename, version, days):
        """Export to Excel with multiple sheets and pass rate chart"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create a summary sheet first
        summary_sheet = wb.create_sheet('Summary', 0)

        # Add version and date range info at the top
        summary_sheet['A1'] = f'Version: {version}'
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet['A2'] = f'Time Range: {days} days'
        summary_sheet['A2'].font = Font(bold=True, size=14)
        summary_sheet['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        summary_sheet['A3'].font = Font(size=11, italic=True)

        # Headers starting from row 5
        summary_sheet['A5'] = 'Platform'
        summary_sheet['B5'] = 'Pass Rate (%)'
        summary_sheet['C5'] = 'Total Tests'
        summary_sheet['D5'] = 'Total Executions'
        summary_sheet['E5'] = 'Passed'
        summary_sheet['F5'] = 'Failed'

        # Style header
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = summary_sheet[f'{col}5']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Add platform summary data (starting from row 6)
        row = 6
        for platform, tests in all_data.items():
            if not tests:
                continue

            total_tests = len(tests)
            total_executions = sum(test['total_runs'] for test in tests)
            passed_executions = sum(test['passed_runs'] for test in tests)
            failed_executions = total_executions - passed_executions
            pass_rate = pass_rates.get(platform, 0)

            summary_sheet[f'A{row}'] = platform
            summary_sheet[f'B{row}'] = round(pass_rate, 1)
            summary_sheet[f'C{row}'] = total_tests
            summary_sheet[f'D{row}'] = total_executions
            summary_sheet[f'E{row}'] = passed_executions
            summary_sheet[f'F{row}'] = failed_executions
            row += 1

        # Add pass rate pie chart
        if len(all_data) > 0:
            chart = PieChart()
            chart.title = 'Pass Rate by Platform'
            chart.height = 12
            chart.width = 20

            labels = Reference(summary_sheet, min_col=1, min_row=6, max_row=row-1)
            data = Reference(summary_sheet, min_col=2, min_row=5, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)

            summary_sheet.add_chart(chart, 'H6')

        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 18
        summary_sheet.column_dimensions['E'].width = 15
        summary_sheet.column_dimensions['F'].width = 15

        # Create Variants sheet
        variants_sheet = wb.create_sheet('Variants')

        # Headers
        variants_sheet['A1'] = 'Platform'
        variants_sheet['B1'] = 'Variant'
        variants_sheet['C1'] = 'Job URL'
        variants_sheet['D1'] = 'Build Date'

        # Style header
        for col in ['A', 'B', 'C', 'D']:
            cell = variants_sheet[f'{col}1']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Query for latest job runs per platform variant
        end_date = datetime.now()
        start_date = end_date - timedelta(days=int(days))

        # Get all unique job runs for this version within time range
        variant_query = """
            SELECT DISTINCT job_name, platform, job_url, timestamp, build_id
            FROM job_runs
            WHERE version = ? AND timestamp >= ? AND timestamp <= ?
            ORDER BY platform, job_name, timestamp DESC
        """
        variant_results = db.execute_query(variant_query, [version, start_date, end_date])

        # Extract variant info from job names
        def extract_variant(job_name, platform):
            """Extract variant name from job name"""
            job_lower = job_name.lower()

            # Check for known variants
            if 'proxy' in job_lower and platform.lower() == 'vsphere':
                return 'proxy'
            elif 'disconnected' in job_lower and platform.lower() == 'vsphere':
                return 'disconnected'
            elif 'upi' in job_lower:
                return 'upi'
            elif 'ipi' in job_lower:
                # Default IPI (not proxy, not disconnected)
                if 'proxy' not in job_lower and 'disconnected' not in job_lower:
                    return 'ipi-connected'

            # Default fallback
            return 'ipi-connected'

        # Group by platform and variant, keep only the latest run for each
        variant_data = {}
        for row in variant_results:
            platform = row['platform']
            job_name = row['job_name']
            variant = extract_variant(job_name, platform)
            key = (platform, variant)

            # Keep only the latest run for each platform-variant combination
            if key not in variant_data:
                variant_data[key] = {
                    'job_url': row['job_url'],
                    'timestamp': row['timestamp'],
                    'job_name': job_name
                }

        # Write variant data to sheet
        row_num = 2
        for (platform, variant), data in sorted(variant_data.items()):
            variants_sheet.cell(row=row_num, column=1, value=platform)
            variants_sheet.cell(row=row_num, column=2, value=variant)

            job_url = data['job_url'] or ''
            variants_sheet.cell(row=row_num, column=3, value=job_url)

            # Make URL clickable
            if job_url:
                cell = variants_sheet.cell(row=row_num, column=3)
                cell.hyperlink = job_url
                cell.font = Font(color='0563C1', underline='single')

            # Format timestamp
            timestamp_str = data['timestamp']
            try:
                if isinstance(timestamp_str, str):
                    ts = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                else:
                    ts = timestamp_str
                build_date = ts.strftime('%Y-%m-%d %H:%M')
            except:
                build_date = str(timestamp_str)

            variants_sheet.cell(row=row_num, column=4, value=build_date)
            row_num += 1

        # Adjust column widths
        variants_sheet.column_dimensions['A'].width = 20
        variants_sheet.column_dimensions['B'].width = 20
        variants_sheet.column_dimensions['C'].width = 80
        variants_sheet.column_dimensions['D'].width = 20

        # Create a sheet for each platform
        for platform, tests in all_data.items():
            if not tests:
                continue

            sheet = wb.create_sheet(platform)

            # Headers
            headers = ['Test ID', 'Title', 'Status', 'Prow URL', 'Comments']
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add test data
            for row_num, test in enumerate(tests, 2):
                # Extract test ID from test_name (e.g., "OCP-12345" or just use the name)
                test_id = test['test_name'].split('-')[0] if '-' in test['test_name'] else test['test_name']

                # Get the most recent NON-SKIPPED run to determine current status
                # We exclude skipped tests from dashboard/export per team policy
                query = """
                    SELECT status, job_url FROM test_results
                    WHERE test_name = ? AND platform = ? AND version = ?
                      AND status != 'skipped'
                    ORDER BY timestamp DESC LIMIT 1
                """
                result = db.execute_query(query, [test['test_name'], platform, version])

                # Determine status and URL from the latest non-skipped run
                job_url = ''
                status = 'Unknown'
                if result and len(result) > 0:
                    latest_status = result[0]['status']
                    job_url = result[0]['job_url'] or ''
                    # Map database status to export status
                    if latest_status == 'passed':
                        status = 'Passed'
                    elif latest_status == 'failed':
                        status = 'Failed'
                    else:
                        status = latest_status.capitalize() if latest_status else 'Unknown'
                else:
                    # If no non-skipped runs found, skip this test entirely
                    continue

                if result and result[0]['job_url']:
                    job_url = result[0]['job_url']

                sheet.cell(row=row_num, column=1, value=test['test_name'])
                sheet.cell(row=row_num, column=2, value=test.get('test_description', ''))
                sheet.cell(row=row_num, column=3, value=status)
                sheet.cell(row=row_num, column=4, value=job_url)
                sheet.cell(row=row_num, column=5, value='')  # Empty comments column

                # Make URL clickable if it exists
                if job_url:
                    cell = sheet.cell(row=row_num, column=4)
                    cell.hyperlink = job_url
                    cell.font = Font(color='0563C1', underline='single')

            # Adjust column widths
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 50
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 60
            sheet.column_dimensions['E'].width = 30

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )

    def export_to_csv(all_data, filename, version, days):
        """Export to CSV with all platforms in one file"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write metadata as comments
        writer.writerow([f'# Version: {version}'])
        writer.writerow([f'# Time Range: {days} days'])
        writer.writerow([f'# Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        writer.writerow([])  # Empty row

        # Write header
        writer.writerow(['Platform', 'Test ID', 'Title', 'Status', 'Prow URL', 'Comments'])

        # Write data for all platforms
        for platform, tests in all_data.items():
            for test in tests:
                test_id = test['test_name']
                title = test.get('test_description', '')

                # Get the most recent NON-SKIPPED run (exclude skipped tests per team policy)
                query = """
                    SELECT status, job_url FROM test_results
                    WHERE test_name = ? AND platform = ? AND version = ?
                      AND status != 'skipped'
                    ORDER BY timestamp DESC LIMIT 1
                """
                result = db.execute_query(query, [test['test_name'], platform, version])

                # Skip tests that only have skipped runs
                if not result or len(result) == 0:
                    continue

                # Determine status and URL from the latest non-skipped run
                job_url = ''
                status = 'Unknown'
                if result and len(result) > 0:
                    latest_status = result[0]['status']
                    job_url = result[0]['job_url'] or ''
                    if latest_status == 'passed':
                        status = 'Passed'
                    elif latest_status == 'failed':
                        status = 'Failed'
                    else:
                        status = latest_status.capitalize() if latest_status else 'Unknown'

                writer.writerow([platform, test_id, title, status, job_url, ''])

        # Convert to bytes
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{filename}.csv'
        )

    def export_to_markdown(all_data, filename, version, days):
        """Export to Markdown with multiple tables"""
        output = io.StringIO()

        output.write(f'# Dashboard Export\n\n')
        output.write(f'**Version:** {version}\n\n')
        output.write(f'**Time Range:** {days} days\n\n')
        output.write(f'**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')

        # Create a table for each platform
        for platform, tests in all_data.items():
            if not tests:
                continue

            # Collect non-skipped tests for this platform
            platform_rows = []
            for test in tests:
                test_id = test['test_name']
                title = test.get('test_description', '')

                # Get the most recent NON-SKIPPED run (exclude skipped tests per team policy)
                query = """
                    SELECT status, job_url FROM test_results
                    WHERE test_name = ? AND platform = ? AND version = ?
                      AND status != 'skipped'
                    ORDER BY timestamp DESC LIMIT 1
                """
                result = db.execute_query(query, [test['test_name'], platform, version])

                # Skip tests that only have skipped runs
                if not result or len(result) == 0:
                    continue

                # Determine status and URL from the latest non-skipped run
                job_url = ''
                status = 'Unknown'
                if result and len(result) > 0:
                    latest_status = result[0]['status']
                    job_url = result[0]['job_url'] or ''
                    if latest_status == 'passed':
                        status = 'Passed'
                    elif latest_status == 'failed':
                        status = 'Failed'
                    else:
                        status = latest_status.capitalize() if latest_status else 'Unknown'

                # Escape pipe characters in title
                title = title.replace('|', '\\|')

                # Format URL as markdown link
                url_display = f'[Link]({job_url})' if job_url else ''

                platform_rows.append(f'| {test_id} | {title} | {status} | {url_display} |\n')

            # Only add platform section if there are non-skipped tests
            if platform_rows:
                output.write(f'## {platform}\n\n')
                output.write('| Test ID | Title | Status | Prow URL |\n')
                output.write('|---------|-------|--------|----------|\n')
                for row in platform_rows:
                    output.write(row)
                output.write('\n')

        # Convert to bytes
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/markdown',
            as_attachment=True,
            download_name=f'{filename}.md'
        )

    @app.teardown_appcontext
    def close_db(error):
        """Close database connection on app shutdown"""
        if error:
            print(f"App error: {error}")

    return app
